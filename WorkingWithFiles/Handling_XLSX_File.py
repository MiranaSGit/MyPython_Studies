from openpyxl import load_workbook
import pandas as pd
import xlrd
import xlsxwriter
newWorkbook = xlsxwriter.Workbook("Stock_Control.xlsx")
newSheet = newWorkbook.add_worksheet("Sheet1")

Part_No = [1, 2, 3, 4, 5]
Stock_Level = [1, 5, 3, 3, 2]

newSheet.write("A1", "Part No")  # Naming Columns in Sheet1
newSheet.write("B1", "Stock Level")

newSheet.write("A2", Part_No[0])  # Putting Values into Columns
newSheet.write("A3", Part_No[1])
newSheet.write("A4", Part_No[2])
newSheet.write("A5", Part_No[3])
newSheet.write("A6", Part_No[4])

newSheet.write("B2", Stock_Level[0])
newSheet.write("B3", Stock_Level[1])
newSheet.write("B4", Stock_Level[2])
newSheet.write("B5", Stock_Level[3])
newSheet.write("B6", Stock_Level[4])

newWorkbook.close()


# Reading from a XLS file

inputWorkbook = xlrd.open_workbook("Stock_Control.xls")
inputWorksheet = inputWorkbook.sheet_by_index(0)  # For opening first sheet

print(inputWorksheet.nrows)  # Listing numbers of row and columns
print(inputWorksheet.ncols)

print(inputWorksheet.cell_value(2, 0))  # Accessing a cell value

Part_no = []
Stock_Level = []

for x in range(inputWorksheet.nrows):
    Part_no.append(inputWorksheet.cell_value(x, 0))
    Stock_Level.append(inputWorksheet.cell_value(x, 1))
print(Part_no, Stock_Level)


# Using Pandas
# import pandas as pd
f = pd.read_excel("smpe.xlsx")

print(f.head())   # Prints first 5 rows
print(f.head(10))   # Prints first 10 rows


print(f.tail(5))  # Prints last 5 rows


# Using Openpyxl
# from openpyxl import load_workbook
f = load_workbook("smpe.xlsx")

print(f.sheetnames)  # Listing sheet names
print(f["smpe"])  # Listing sheet name "smpe"
print(f.active)  # Listing active sheet name

sheet = f.active

new_data = [ABC123, "OFD", 0, 0, 0, 22]

for row in sheet["A5":"G5"]:
    for cell in row:
        print(cell.value)


for row in sheet["A5":"G5"]:
    for index, cell in enumerate(row):
        cell.value = new_data[index]

f.save("new_smpe.xlsx")


# Printing Sheet Names
f2 = load_workbook("SMPE_LW_Stock.xlsx")
print(f2.sheetnames)

sheet2 = f2['Sheet1']  # Setting Sheet name

for cell in sheet2['A']:  # Printing Column A
    if cell.value != None:
        print(cell.value)

f2.close()  # Close the workbook
