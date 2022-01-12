import os
import csv
from openpyxl import load_workbook

abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)

listEBAY = []
dictNAT = {}

for i in os.listdir("."):
    if i.startswith("NATIONAL"):
        f = load_workbook(i)
        sheet = f.active
        for (a, b, c) in sheet.iter_rows(max_col=3):
            if a.value != None and a.value != "NAP Part number" and a.value.startswith("CK"):
                dictNAT[a.value] = c.value
        f.close()
    if i.startswith("eBay-active"):
        with open(i, 'r', newline='', encoding='utf-8') as file:
            csv_rows = csv.reader(file)  # reader() function takes each
            # row (lines) into a list
            header = next(csv_rows)
            if header != None:   # Passing header row
                for row in csv_rows:
                    if row[3].startswith("NT") and row[6].startswith("GBP"):
                        listEBAY.append(row[3])

listnewEBAY = []
for i in listEBAY:
    if 'CK' in i:
        j = i.index('CK')
        listnewEBAY.append(i[j:])

dictNew = {}
with open('TPde_Olmayan_NAT_Items.csv', mode='w', newline="") as new_file:
    for i, j in dictNAT.items():
        if i not in listnewEBAY:
            dictNew[i] = j
    writer = csv.writer(new_file, delimiter=',')
    writer.writerows(dictNew.items())
